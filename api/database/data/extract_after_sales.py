"""
Flattens the two After-Sales workbooks into one JSON the ERP can read.

    MSR  — the Google Form repair-request responses (the intake queue)
    MRR  — the monthly revenue report (what each job cost and earned)

Both are hand-kept spreadsheets, so the job here is as much cleaning as
reading. Three problems are fixed on the way through:

  * Free-text where a list belonged. "Client Classification" holds 279 distinct
    spellings of about six real values; equipment holds 283 spellings of maybe
    thirty. Both are mapped onto a controlled vocabulary, and anything that
    does not map is reported rather than silently bucketed.

  * Columns typed into the wrong column. The MRR's TECHNICIAN column contains
    money in 300-odd rows and its REPAIR TYPE column contains technician names,
    because the two sit next to each other. Values are validated by shape, not
    by position, and a number found under TECHNICIAN is not treated as a name.

  * Totals that disagree with their parts. TOTAL is keyed by hand, so it is
    recomputed from the cost and revenue columns and the difference reported.

    python database/data/extract_after_sales.py MSR.xlsx MRR.xlsx
"""

import json
import os
import re
import sys
from datetime import date, datetime

import openpyxl

# --------------------------------------------------------------------------
# Controlled vocabularies — the values the ERP will actually offer.
# --------------------------------------------------------------------------

CLIENT_TYPES = ["Panadero", "Institutional", "CHBC", "JBYL Group", "PDF", "Company-Owned", "Franchise"]

CLIENT_TYPE_PATTERNS = [
    (r"PANADERO", "Panadero"),
    (r"\bCHBC\b|CHEFS?\s*HUT", "CHBC"),
    (r"\bJBYL\b", "JBYL Group"),
    (r"\bPDF\b", "PDF"),
    (r"COMPANY\s*[-–]?\s*OWNED|\bPKE\b", "Company-Owned"),
    (r"FRANCHIS", "Franchise"),
    (r"INSTITUTION|GENERAL", "Institutional"),
]

EQUIPMENT_PATTERNS = [
    (r"SPIRAL\s*MIX", "Spiral Mixer"),
    (r"PLANETARY\s*MIX", "Planetary Mixer"),
    (r"\bMIXER\b", "Mixer"),
    (r"DECK\s*OVEN|\bOVEN\b|ROTARY", "Oven"),
    (r"CAKE\s*CHILL", "Cake Chiller"),
    (r"CHILL", "Chiller"),
    (r"FREEZ", "Freezer"),
    (r"\bFRYER\b", "Fryer"),
    (r"RICE\s*STEAM|STEAMER", "Rice Steamer"),
    (r"DISH\s*WASH", "Dish Washer"),
    (r"PROOFER", "Proofer"),
    (r"ROTISSER", "Rotisserie"),
    (r"BREAD\s*SHOWCASE|SHOWCASE", "Bread Showcase"),
    (r"MEAT\s*SLICER", "Meat Slicer"),
    (r"BREAD\s*SLICER|SLICER", "Bread Slicer"),
    (r"ICE\s*MAKER", "Ice Maker"),
    (r"CHINESE\s*WOK|\bWOK\b", "Chinese Wok"),
    (r"GRIDDLE|GRILL|ROBATA", "Griddle / Grill"),
    (r"\bAIRCON|\bACU\b|SPLIT\s*TYPE", "Air Conditioning"),
    (r"DOUGH\s*SHEET", "Dough Sheeter"),
    (r"WATER\s*DISP|DISPENSER", "Water Dispenser"),
    (r"EXHAUST|HOOD|BLOWER", "Exhaust / Hood"),
    (r"\bSTOVE\b|BURNER|RANGE", "Stove / Range"),
]

REPAIR_TYPES = [
    "Check-up",
    "Minor Repair",
    "Major Repair",
    "PMS",
    "System Reprocess",
    "Commissioning",
    "Installation",
    "Others",
]

REPAIR_TYPE_PATTERNS = [
    (r"CHECK\s*-?\s*UP|CHECKUP", "Check-up"),
    (r"MINOR", "Minor Repair"),
    (r"MAJOR", "Major Repair"),
    (r"\bPMS\b|PREVENTIVE|PREDICTIVE", "PMS"),
    (r"REPROCESS", "System Reprocess"),
    (r"COMMISSION", "Commissioning"),
    (r"INSTALL", "Installation"),
    (r"DELIVER|DEMO|HAUL", "Others"),
]

# Status on the MSR sheet.
STATUS_MAP = {
    "DONE": "Done",
    "CANCELED": "Cancelled",
    "CANCELLED": "Cancelled",
    "FOR SCHEDULE": "For Scheduling",
    "SCHEDULED": "Scheduled",
    "PENDING": "Pending",
}

# The MRR's money columns, by header name.
COST_COLUMNS = ["FUEL", "BARGE", "ACCOMMODITION", "MEALS", "TRANSPORTATION"]
REVENUE_COLUMNS = [
    "CHECK - UP",
    "MINOR REPAIR",
    "MAJOR REPAIR",
    "PMS",
    "SYSTEM REPROCESS",
    "COMMISSIONING",
    "INSTALLATION",
    "OTHERS",
]


def norm(value) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def match(value: str, patterns, fallback=None):
    upper = norm(value).upper()
    if not upper:
        return fallback
    for pattern, label in patterns:
        if re.search(pattern, upper):
            return label
    return fallback


def money(value) -> float:
    """A money cell that holds a name is not money."""
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[^\d.\-]", "", str(value))
    try:
        return float(cleaned) if cleaned not in ("", "-", ".") else 0.0
    except ValueError:
        return 0.0


def looks_like_name(value) -> bool:
    """Guards the TECHNICIAN column, which is full of stray money values."""
    text = norm(value)
    return bool(text) and not re.fullmatch(r"[\d.,\s-]+", text)


def iso(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return None


def priority(value) -> int | None:
    found = re.search(r"PRIORITY\s*([1-4])", norm(value).upper())
    return int(found.group(1)) if found else None


# --------------------------------------------------------------------------


def read_msr(path: str, report: dict) -> list[dict]:
    sheet = openpyxl.load_workbook(path, data_only=True)["Form Responses 1"]
    header = [norm(sheet.cell(1, c).value) for c in range(1, sheet.max_column + 1)]
    index = {name: i for i, name in enumerate(header)}

    def cell(row, name):
        i = index.get(name)
        return row[i] if i is not None and i < len(row) else None

    out = []
    unmapped_equipment = set()
    unmapped_client = set()

    for r in range(2, sheet.max_row + 1):
        row = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
        if not any(v not in (None, "") for v in row):
            continue

        ticket = norm(cell(row, "Repair Ticket Number"))
        if not ticket:
            continue

        raw_equipment = norm(cell(row, "Type of Equipment (Please choose one equipment only)"))
        raw_client = norm(cell(row, "Client Classification"))

        equipment = match(raw_equipment, EQUIPMENT_PATTERNS)
        if raw_equipment and not equipment:
            unmapped_equipment.add(raw_equipment)

        client_type = match(raw_client, CLIENT_TYPE_PATTERNS)
        if raw_client and not client_type:
            unmapped_client.add(raw_client)

        out.append(
            {
                "ticket": re.sub(r"^RT\s*#\s*", "", ticket).strip(),
                "requestedAt": iso(cell(row, "Timestamp")),
                "status": STATUS_MAP.get(norm(cell(row, "STATUS")).upper(), "Pending"),
                "remarks": norm(cell(row, "REMARKS")),
                "client": norm(cell(row, "Business / Trade Name")),
                "branch": norm(cell(row, "Branch / Address")),
                "clientType": client_type or "Institutional",
                "clientTypeRaw": raw_client,
                "contact": norm(cell(row, "Contact person")),
                "phone": norm(cell(row, "Contact number")),
                "email": norm(cell(row, "Email Address")),
                "preferredTime": norm(cell(row, "Preferred Time of the Day for Repair")),
                "requestType": norm(cell(row, "Type of Request")) or "New Repair Request",
                "priority": priority(cell(row, "Urgency/Priority")),
                "equipment": equipment or "Others",
                "equipmentRaw": raw_equipment,
                "issue": norm(cell(row, "Equipment Concern / Issue Description")),
                "attachment": norm(cell(row, "Attach Supporting Photos / Video")),
            }
        )

    report["msr_unmapped_equipment"] = sorted(unmapped_equipment)[:40]
    report["msr_unmapped_client"] = sorted(unmapped_client)[:40]
    return out


def read_mrr(path: str, report: dict) -> list[dict]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    out = []
    total_mismatch = 0

    for name in workbook.sheetnames:
        sheet = workbook[name]
        header = [norm(sheet.cell(1, c).value).upper().replace("\n", " ") for c in range(1, sheet.max_column + 1)]
        index = {h: i for i, h in enumerate(header)}

        def cell(row, key):
            i = index.get(key)
            return row[i] if i is not None and i < len(row) else None

        for r in range(2, sheet.max_row + 1):
            row = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
            tsr = norm(cell(row, "TSR"))
            client = norm(cell(row, "CLIENT'S NAME"))
            if not tsr and not client:
                continue

            costs = {c.title(): money(cell(row, c)) for c in COST_COLUMNS}
            revenue = {c.title(): money(cell(row, c)) for c in REVENUE_COLUMNS}

            cost_total = sum(costs.values())
            revenue_total = sum(revenue.values())
            stated = money(cell(row, "TOTAL"))
            if stated and abs(stated - (cost_total + revenue_total)) > 0.5:
                total_mismatch += 1

            technician = cell(row, "TECHNICIAN")
            # The two REPAIR TYPE columns are the request and its classification;
            # openpyxl keeps only the last of a duplicated header, so both are
            # read positionally instead.
            repair_cols = [i for i, h in enumerate(header) if h == "REPAIR TYPE"]
            requested = norm(row[repair_cols[0]]) if repair_cols else ""
            classified = norm(row[repair_cols[1]]) if len(repair_cols) > 1 else ""

            out.append(
                {
                    "sheet": name,
                    "tsr": tsr,
                    "ticket": norm(cell(row, "REPAIR TICKET #")),
                    "repairedOn": iso(cell(row, "DATE OF REPAIR")),
                    "submittedOn": iso(cell(row, "DATE SUBMITTED")),
                    "clientType": match(cell(row, "CLIENT TYPE"), CLIENT_TYPE_PATTERNS) or "Institutional",
                    "client": client,
                    "address": norm(cell(row, "CLIENT'S ADDRESS")),
                    "equipment": match(cell(row, "EQUIPMENT"), EQUIPMENT_PATTERNS) or "Others",
                    "equipmentRaw": norm(cell(row, "EQUIPMENT")),
                    "srNo": norm(cell(row, "SR #")),
                    "drNo": norm(cell(row, "DR #")),
                    "requestedWork": requested,
                    # Classification is only trusted when it maps to a real
                    # repair type; in ~90 rows a technician's name was typed here.
                    "repairType": match(classified, REPAIR_TYPE_PATTERNS)
                    or match(requested, REPAIR_TYPE_PATTERNS)
                    or "Others",
                    "technicians": [t.strip() for t in re.split(r"[|,/&]", norm(technician)) if t.strip()]
                    if looks_like_name(technician)
                    else [],
                    "costs": costs,
                    "revenue": revenue,
                    "costTotal": round(cost_total, 2),
                    "revenueTotal": round(revenue_total, 2),
                    "statedTotal": stated,
                }
            )

    report["mrr_total_mismatches"] = total_mismatch
    return out


def main(msr_path: str, mrr_path: str) -> None:
    report: dict = {}
    msr = read_msr(msr_path, report)
    mrr = read_mrr(mrr_path, report)

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "after_sales.json")
    with open(out, "w", encoding="utf-8") as handle:
        json.dump(
            {
                "vocabularies": {
                    "clientTypes": CLIENT_TYPES,
                    "repairTypes": REPAIR_TYPES,
                    "equipment": sorted({label for _, label in EQUIPMENT_PATTERNS}),
                },
                "requests": msr,
                "jobs": mrr,
            },
            handle,
            indent=1,
        )

    billed = [j for j in mrr if j["revenueTotal"] > 0]
    print(f"service requests : {len(msr)}")
    print(f"revenue jobs     : {len(mrr)}  ({len(billed)} with revenue)")
    print(f"revenue captured : {sum(j['revenueTotal'] for j in mrr):,.2f}")
    print(f"costs captured   : {sum(j['costTotal'] for j in mrr):,.2f}")
    print(f"TOTAL disagreeing with its parts: {report['mrr_total_mismatches']} row(s)")
    if report["msr_unmapped_equipment"]:
        print(f"unmapped equipment ({len(report['msr_unmapped_equipment'])}): {report['msr_unmapped_equipment'][:6]}")
    if report["msr_unmapped_client"]:
        print(f"unmapped client types ({len(report['msr_unmapped_client'])}): {report['msr_unmapped_client'][:6]}")
    print(f"written to: {out}")


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
