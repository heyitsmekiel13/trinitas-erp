"""
Flattens the QuickBooks inventory export into a CSV the importer can read.

The export is a hierarchy expressed by indentation: the item's name sits in
whichever column matches its depth, and the group headings above it carry the
category. Rows whose label starts with "Total" are QuickBooks' own subtotals and
are dropped — re-importing them would double every quantity.

Run once when a fresh export arrives:

    python database/data/extract_qb_inventory.py "path/to/export.xlsx"

Writes database/data/qb_inventory.csv next to this script.
"""

import csv
import os
import re
import sys

import openpyxl

SHOWROOM_COL = 6   # F
WAREHOUSE_COL = 8  # H
LABEL_COLS = range(1, 6)  # A..E — the indentation carries the depth
CATEGORY_COL = 3          # C — where QuickBooks' real categories sit


def clean(name: str) -> tuple[str, str]:
    """
    QuickBooks writes "Item Name (Description)".

    The two are identical for most rows, so the parenthetical is kept only when
    it actually says something the name does not.

    A long label is truncated by QuickBooks *mid-description*, leaving an
    unterminated bracket and an ellipsis — "1 Lot Malandag Hardwa2 (1 Lot
    Malandag Hardware Materials (2nd ...". Those still split cleanly at the
    first bracket; taking the whole string as the name would carry the ellipsis
    into the catalogue.
    """
    name = re.sub(r"\s+", " ", str(name)).strip()

    match = re.match(r"^(.*?)\s*\((.*)\)$", name)
    if not match:
        # Unterminated: split at the first bracket and drop the ellipsis.
        if "(" in name:
            label, _, rest = name.partition("(")
            label = label.strip()
            description = rest.strip().rstrip(".").strip()
            if label:
                return label, "" if description.lower() == label.lower() else description
        return name, ""

    label, description = match.group(1).strip(), match.group(2).strip()
    if not label:
        return description, ""
    if description.lower() == label.lower():
        return label, ""
    return label, description


def qty(value) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def main(source: str) -> None:
    workbook = openpyxl.load_workbook(source, data_only=True)
    sheet = workbook["Sheet1"]

    rows = []
    # Depth -> the group heading currently open at that depth.
    open_groups: dict[int, str] = {}
    skipped_totals = 0

    for row in range(1, sheet.max_row + 1):
        label = None
        depth = None
        for column in LABEL_COLS:
            value = sheet.cell(row, column).value
            if value is not None and str(value).strip():
                label = str(value).strip()
                depth = column
                break

        if label is None:
            continue

        # Subtotals close a group; they are never items.
        if label.lower().startswith("total"):
            skipped_totals += 1
            open_groups.pop(depth, None)
            continue

        showroom = sheet.cell(row, SHOWROOM_COL).value
        warehouse = sheet.cell(row, WAREHOUSE_COL).value

        # A heading has no quantities of its own — the subtotal row carries them.
        if showroom is None and warehouse is None:
            open_groups[depth] = clean(label)[0]
            # Anything nested deeper belonged to a heading that just closed.
            for deeper in [d for d in open_groups if d > depth]:
                open_groups.pop(deeper)
            continue

        name, description = clean(label)

        # QuickBooks' real categories all sit at one depth (column C). Anything
        # deeper is a sub-item of a parent *item*, not a new category — taking
        # the nearest heading would turn a parent item's name into a category of
        # its own, which is how "Chopstick Stainless" became a category.
        category = open_groups.get(CATEGORY_COL, "Uncategorised")
        if category == "Inventory":
            category = "Uncategorised"

        # A sub-item's parent is worth keeping; it is the only thing that
        # distinguishes it from its siblings once the indentation is gone.
        parent = open_groups.get(depth - 1)
        if parent and parent != category and depth > CATEGORY_COL + 1:
            description = f"{parent} — {description}" if description else parent

        rows.append(
            {
                "category": category,
                "name": name,
                "description": description,
                "showroom": qty(showroom),
                "warehouse": qty(warehouse),
                "source_row": row,
            }
        )

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "qb_inventory.csv")
    with open(out, "w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle, fieldnames=["category", "name", "description", "showroom", "warehouse", "source_row"]
        )
        writer.writeheader()
        writer.writerows(rows)

    total_showroom = sum(r["showroom"] for r in rows)
    total_warehouse = sum(r["warehouse"] for r in rows)
    print(f"items:      {len(rows)}")
    print(f"subtotals skipped: {skipped_totals}")
    print(f"showroom:   {total_showroom:,.2f}")
    print(f"warehouse:  {total_warehouse:,.2f}")
    print(f"total:      {total_showroom + total_warehouse:,.2f}")
    print(f"written to: {out}")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "QB Inventory.xlsx")
